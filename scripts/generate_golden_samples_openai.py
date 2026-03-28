#!/usr/bin/env python3
"""
Sinh thêm dòng golden intent bằng GPT-4o (OpenAI API), format giống sheet Dataset
trong data/golden_intent_dataset.xlsx: User Query + Lý luận (VI) + Label.

KEY: đặt OPENAI_API_KEY trong file .env ở root repo (KHÔNG commit key).

Chạy từ root:
  pip install openai python-dotenv openpyxl
  python scripts/generate_golden_samples_openai.py
  python scripts/generate_golden_samples_openai.py --total 200
  python scripts/generate_golden_samples_openai.py --per-label 30 --out data/golden_intent_gpt4o.xlsx

Mặc định --total 230 (≈39+39+38+38+38+38). Nếu file chỉ có 150 dòng = thường do đã chạy
  --total 150 hoặc --per-label 25 — chạy lại không thêm flag để ghi đè đủ 230 mẫu.

Tham khảo prompt: hằng SYSTEM_PROMPT_INTENT và build_user_message() bên dưới.
"""
from __future__ import annotations

import argparse
import importlib.util
import json
import os
import sys
from pathlib import Path

from dotenv import load_dotenv
from openai import OpenAI
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]


def _balanced_per_label(total: int, k: int) -> list[int]:
    """Chia `total` thành k phần số nguyên, sai lệch tối đa 1 (phần dư vào các nhãn đầu)."""
    if k <= 0:
        return []
    base = total // k
    rem = total % k
    return [base + (1 if i < rem else 0) for i in range(k)]


def _ensure_utf8_stdout() -> None:
    w = getattr(sys.stdout, "reconfigure", None)
    if callable(w):
        try:
            sys.stdout.reconfigure(encoding="utf-8")
        except Exception:
            pass


# Load taxonomy từ script build gốc (một nguồn sự thật)
_spec = importlib.util.spec_from_file_location(
    "build_golden_intent_dataset",
    ROOT / "scripts" / "build_golden_intent_dataset.py",
)
_mod = importlib.util.module_from_spec(_spec)
assert _spec.loader is not None
_spec.loader.exec_module(_mod)
TAXONOMY: list[tuple[str, str]] = _mod.TAXONOMY
LABELS = [t[0] for t in TAXONOMY]

# --- Prompt (có thể chỉnh trực tiếp tại đây) ---

SYSTEM_PROMPT_INTENT = """Bạn là người tạo dữ liệu huấn luyện cho bộ phân loại ý định (intent) của chatbot đồng hành anime/game.

Nhiệm vụ: sinh các câu người dùng GIẢ ĐỊNH (không phải lời khuyên y tế thật) để mô hình học NHẬN DIỆN nhãn.

Quy tắc chung:
- Mỗi câu phải khác nhau về chủ đề, ngữ điệu, độ dài; tránh lặp template.
- Câu thường DÀI (nhiều câu nhỏ), tự nhiên như chat: có thể lẫn tiếng Việt và tiếng Anh trong cùng một đoạn (code-switching), giống người trẻ nhắn tin.
- "Lý luận gợi ý" viết bằng tiếng Việt, ngắn gọn, giải thích vì sao câu đó khớp ĐÚNG nhãn được giao (1–3 câu).
- Nhãn (label) phải là một trong các mã được cung cấp, đúng chính tả snake_case.
- Với nhãn crisis_alert: chỉ mô tả tín hiệu tâm lý / ý nghĩ tự hại để phân loại an toàn — không hướng dẫn phương pháp tự hại, không cổ vũ. Giữ giọng điệu như người đang cần hỗ trợ khẩn."""

USER_PROMPT_TEMPLATE = """Bạn cần sinh ĐÚNG {n} mẫu cho MỘT nhãn duy nhất.

Nhãn (label): `{label}`
Mô tả nhãn: {description}

Yêu cầu:
- Đủ {n} object trong mảng `samples`.
- Mỗi object có khóa: "user_query" (string), "reasoning_vi" (string tiếng Việt).
- user_query: đoạn người dùng giả định, dài, đa dạng; có thể có tiếng Anh xen kẽ.
- reasoning_vi: giải thích vì sao gán nhãn `{label}`.

Trả về DUY NHẤT một JSON object hợp lệ (không markdown), dạng:
{{"samples":[{{"user_query":"...","reasoning_vi":"..."}},...]}}"""

# Mỗi lần gọi API chỉ xin tối đa nhiêu mẫu — tránh JSON bị cắt (output limit) khi xin 30–40 mẫu/lần.
MAX_SAMPLES_PER_API_CALL = 18
MAX_CHUNK_RETRIES = 4


def build_user_message(
    label: str,
    description: str,
    n: int,
    *,
    continuation: str = "",
) -> str:
    base = USER_PROMPT_TEMPLATE.format(
        n=n,
        label=label,
        description=description,
    )
    if continuation:
        return base + "\n\n" + continuation
    return base


def _parse_json_object(content: str) -> dict:
    content = content.strip()
    if content.startswith("```"):
        lines = content.split("\n")
        lines = lines[1:]
        if lines and lines[-1].strip() == "```":
            lines = lines[:-1]
        content = "\n".join(lines)
    return json.loads(content)


def _style_header(ws, row: int = 1) -> None:
    for cell in ws[row]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def _autosize_columns(ws, max_width: int = 72) -> None:
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_len = 0
        for row in ws.iter_rows(min_col=col, max_col=col):
            for c in row:
                if c.value is not None:
                    max_len = max(max_len, min(120, len(str(c.value))))
        ws.column_dimensions[letter].width = min(max_width, max(14, max_len + 2))


def _fallback_crisis(n: int) -> list[dict[str, str]]:
    """Nếu API từ chối block crisis, dùng mẫu ngắn (vẫn đủ để train classifier)."""
    templates = [
        (
            "I don't see the point anymore — everything feels empty and I keep thinking people would be better off without me.",
            "Tín hiệu tuyệt vọng và ý nghĩ không muốn tồn tại — crisis_alert.",
        ),
        (
            "Đêm nay mình run sợ vì trong đầu cứ lặp ý nghĩ làm hại bản thân; không dám nói với ai ở nhà.",
            "Có ý tưởng tự hại và sợ hãi — crisis_alert.",
        ),
        (
            "I've been stockpiling pills 'just in case' — I know it's wrong but the impulse scares me.",
            "Chuẩn bị phương tiện / ý định nguy hiểm — crisis_alert.",
        ),
    ]
    out: list[dict[str, str]] = []
    for i in range(n):
        u, r = templates[i % len(templates)]
        out.append({"user_query": f"{u} [variant {i}]", "reasoning_vi": r})
    return out


def _one_api_chunk(
    client: OpenAI,
    model: str,
    label: str,
    description: str,
    n: int,
    *,
    already: int,
    batch_idx: int,
) -> list[dict[str, str]]:
    """Một lần gọi API, yêu cầu đúng `n` mẫu (n nhỏ, thường ≤ MAX_SAMPLES_PER_API_CALL)."""
    cont = ""
    if already > 0:
        cont = (
            f"(Phần tiếp theo — batch #{batch_idx + 1}: đã có {already} mẫu trước đó cho nhãn `{label}`. "
            f"Sinh thêm ĐÚNG {n} mẫu MỚI, không lặp lại nội dung các mẫu trước.)"
        )
    user_msg = build_user_message(label, description, n, continuation=cont)
    resp = client.chat.completions.create(
        model=model,
        temperature=0.85,
        max_tokens=16384,
        response_format={"type": "json_object"},
        messages=[
            {"role": "system", "content": SYSTEM_PROMPT_INTENT},
            {"role": "user", "content": user_msg},
        ],
    )
    raw = resp.choices[0].message.content or "{}"
    data = _parse_json_object(raw)
    samples = data.get("samples")
    if not isinstance(samples, list):
        raise ValueError(f"Missing samples array for label={label}")
    cleaned: list[dict[str, str]] = []
    for item in samples:
        if not isinstance(item, dict):
            continue
        uq = str(item.get("user_query", "")).strip()
        rv = str(item.get("reasoning_vi", "")).strip()
        if uq and rv:
            cleaned.append({"user_query": uq, "reasoning_vi": rv})
    return cleaned[:n]


def generate_for_label(
    client: OpenAI,
    model: str,
    label: str,
    description: str,
    n: int,
) -> list[dict[str, str]]:
    if n <= 0:
        return []
    out: list[dict[str, str]] = []
    batch_idx = 0
    while len(out) < n:
        need = n - len(out)
        chunk_target = min(MAX_SAMPLES_PER_API_CALL, need)
        got: list[dict[str, str]] = []
        last_err: Exception | None = None
        for attempt in range(MAX_CHUNK_RETRIES):
            try:
                got = _one_api_chunk(
                    client,
                    model,
                    label,
                    description,
                    chunk_target,
                    already=len(out),
                    batch_idx=batch_idx,
                )
            except Exception as e:
                last_err = e
                got = []
            if len(got) >= chunk_target:
                break
        if len(got) < chunk_target:
            if label == "crisis_alert":
                out.extend(got)
                still = n - len(out)
                if still > 0:
                    out.extend(_fallback_crisis(still)[:still])
                break
            raise RuntimeError(
                f"Sau {MAX_CHUNK_RETRIES} lần thử vẫn thiếu mẫu cho {label}: "
                f"cần {chunk_target}, nhận {len(got)}. Lỗi cuối: {last_err}"
            )
        out.extend(got[:chunk_target])
        batch_idx += 1
    if len(out) != n:
        raise RuntimeError(f"Internal: label={label} expected {n} samples, got {len(out)}")
    return out[:n]


def write_xlsx(
    path: Path,
    rows: list[tuple[str, str, str]],
    note: str,
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws_t = wb.active
    ws_t.title = "Taxonomy"
    ws_t.append(["label", "description"])
    for lab, desc in TAXONOMY:
        ws_t.append([lab, desc])
    _style_header(ws_t)
    _autosize_columns(ws_t)

    ws_d = wb.create_sheet("Dataset")
    ws_d.append(
        [
            "stt",
            "Câu người dùng (User Query)",
            "Lý luận gợi ý (Reasoning, tiếng Việt)",
            "Nhãn (Label)",
            "ghi_chú",
        ]
    )
    for stt, (uq, rv, lab) in enumerate(rows, start=1):
        ws_d.append([stt, uq, rv, lab, note])
    _style_header(ws_d)
    _autosize_columns(ws_d)
    wb.save(path)


def main() -> int:
    _ensure_utf8_stdout()
    load_dotenv(ROOT / ".env")
    api_key = os.environ.get("OPENAI_API_KEY", "").strip()
    if not api_key or api_key.startswith("sk-your"):
        print(
            "Thiếu OPENAI_API_KEY trong .env (hoặc vẫn là placeholder). "
            "Đặt key thật rồi chạy lại — không hardcode trong repo.",
            file=sys.stderr,
        )
        return 1

    p = argparse.ArgumentParser(description="Sinh golden samples qua GPT-4o")
    p.add_argument(
        "--total",
        type=int,
        default=230,
        help=(
            "Tổng số mẫu, chia đều 6 nhãn (±1). Mặc định 230 — bù 50 mẫu so với lần chạy cũ (30×6=180)."
        ),
    )
    p.add_argument(
        "--per-label",
        type=int,
        default=None,
        metavar="N",
        help="Nếu đặt: mỗi nhãn đúng N mẫu (ghi đè --total). Ví dụ 30 → 180 mẫu.",
    )
    p.add_argument("--model", default="gpt-4o", help="Tên model OpenAI")
    p.add_argument(
        "--out",
        type=Path,
        default=ROOT / "data" / "golden_intent_gpt4o.xlsx",
        help="File Excel output",
    )
    p.add_argument(
        "--dry-run",
        action="store_true",
        help="Chỉ in system prompt + ví dụ user message rồi thoát",
    )
    args = p.parse_args()

    if args.per_label is not None:
        counts = [args.per_label] * len(TAXONOMY)
    else:
        counts = _balanced_per_label(args.total, len(TAXONOMY))

    if args.dry_run:
        print("=== SYSTEM (rút gọn) ===\n", SYSTEM_PROMPT_INTENT[:800], "...\n")
        lab, desc = TAXONOMY[0]
        n0 = counts[0]
        print("=== Phân bổ mẫu / nhãn ===")
        print(dict(zip([t[0] for t in TAXONOMY], counts)), "→ tổng", sum(counts))
        print("=== USER (ví dụ nhãn đầu) ===\n")
        print(build_user_message(lab, desc, n0))
        return 0

    client = OpenAI(api_key=api_key)
    all_rows: list[tuple[str, str, str]] = []

    for (label, description), n in zip(TAXONOMY, counts):
        print(f"Generating {n} samples for {label}...", flush=True)
        try:
            batch = generate_for_label(client, args.model, label, description, n)
        except Exception as e:
            if label == "crisis_alert":
                print(f"API issue on crisis_alert ({e}), using fallback.", flush=True)
                batch = _fallback_crisis(n)
            else:
                raise
        for item in batch:
            all_rows.append(
                (item["user_query"], item["reasoning_vi"], label),
            )

    expected_n = sum(counts)
    if len(all_rows) != expected_n:
        raise RuntimeError(f"Số dòng không khớp: có {len(all_rows)}, cần {expected_n}")
    write_xlsx(args.out, all_rows, note=f"gpt-4o:{args.model}")
    print(
        f"Wrote {args.out} — {len(all_rows)} rows "
        f"(per-label counts: {list(zip([t[0] for t in TAXONOMY], counts))})."
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
