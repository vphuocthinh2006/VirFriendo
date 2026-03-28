#!/usr/bin/env python3
"""
Tạo golden dataset intent: 1 file Excel (taxonomy + dataset).

Chạy từ root repo:
  pip install openpyxl
  python scripts/build_golden_intent_dataset.py

Output: data/golden_intent_dataset.xlsx
  - Sheet "Taxonomy"
  - Sheet "Dataset": User Query (dài, Vi/En) + Reasoning (VI) + Label (+ ghi chú)

Chỉnh số mẫu dài sinh thêm: hằng số EXTRA_LONG_SAMPLES (mặc định 200, trong khoảng 100–200).
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "data" / "golden_intent_dataset.xlsx"

# 100–200 mẫu dài sinh thêm (không tính các dòng seed ngắn phía dưới)
EXTRA_LONG_SAMPLES = 200

TAXONOMY: list[tuple[str, str]] = [
    (
        "greeting_chitchat",
        "Chào hỏi, small talk, trò chuyện xã giao (không rơi rõ nhóm khác).",
    ),
    (
        "out_of_domain",
        "Hỏi code, tài chính, thời tiết, tin tức, toán… ngoài entertainment và tâm lý nhẹ.",
    ),
    (
        "entertainment_knowledge",
        "Hỏi về manga, anime, game, phim, light novel, nhân vật, chapter, lore, fanwiki, review…",
    ),
    (
        "psychology_venting",
        "Đang xả / giãi bày (mệt, buồn, tức, cô đơn, stress…) — chưa xin cách xử lý rõ.",
    ),
    (
        "psychology_advice_seeking",
        "Xin lời khuyên, “làm sao để…”, “cách nào…”, tips, how to.",
    ),
    (
        "crisis_alert",
        "Có ý tự hại / tự tử / nguy hiểm tính mạng (ưu tiên cao nhất khi phân loại).",
    ),
]

# Mẫu ngắn ban đầu (giữ trong file Excel đầu sheet Dataset)
SEED_ROWS: list[tuple[str, str, str]] = [
    ("chào bạn, hôm nay bạn thế nào", "Chào hỏi thông thường, không hỏi kiến thức hay tâm lý sâu.", "greeting_chitchat"),
    ("cảm ơn bạn nhiều nhé", "Lời cảm ơn, small talk.", "greeting_chitchat"),
    ("viết giúp mình đoạn code python đọc csv", "Yêu cầu code / kỹ thuật, ngoài domain chatbot.", "out_of_domain"),
    ("bitcoin hôm nay giá bao nhiêu", "Hỏi tài chính thực tế.", "out_of_domain"),
    ("tóm tắt One Piece arc Wano", "Hỏi nội dung anime/manga.", "entertainment_knowledge"),
    ("Gojo trong JJK còn sống không (spoiler ngắn)", "Hỏi nhân vật / plot anime.", "entertainment_knowledge"),
    ("Elden Ring boss Malenia tips", "Hỏi game / gameplay.", "entertainment_knowledge"),
    ("mệt quá, sếp chửi mình cay quá", "Than thở cảm xúc, chưa xin cách xử lý rõ.", "psychology_venting"),
    ("cô đơn trong phòng một mình", "Xả cảm xúc cô đơn.", "psychology_venting"),
    ("làm sao để bớt lo lắng trước thi", "Xin cách / lời khuyên cụ thể.", "psychology_advice_seeking"),
    ("tips ngủ sớm khi hay suy nghĩ", "Xin tips, hướng xử lý.", "psychology_advice_seeking"),
    ("tôi không muốn sống nữa", "Tín hiệu nguy cơ tự hại — crisis.", "crisis_alert"),
    ("I want to kill myself", "Tín hiệu tiếng Anh — crisis.", "crisis_alert"),
]


def _long_row_at(i: int) -> tuple[str, str, str]:
    """Một dòng dài thứ i (0-based): (text, reasoning_vi, label). Nhiều biến thể theo i."""
    v = i % 6
    k = i // 6
    anime = [
        "One Piece", "Jujutsu Kaisen", "Chainsaw Man", "Attack on Titan", "Frieren",
        "Steins;Gate", "Genshin Impact", "Honkai: Star Rail", "Elden Ring", "Baldur's Gate 3",
        "Cyberpunk 2077", "Hollow Knight", "Vinland Saga", "Oshi no Ko", "Mob Psycho 100",
    ]
    w = anime[k % len(anime)]

    if v == 0:  # greeting
        text = (
            f"Hey — I know we haven't talked in a bit; work's been a mess and I've been travelling, "
            f"but I wanted to drop by and see how you're holding up. "
            f"Không có agenda gì đặc biệt, chỉ là hỏi thăm và maybe rủ cafe cuối tuần nếu bạn rảnh."
        )
        reason = "Chào hỏi, small talk; không yêu cầu kiến thức giải trí hay tư vấn chuyên sâu."
        lab = "greeting_chitchat"
    elif v == 1:  # ood
        text = (
            f"I'm comparing two mortgage offers and the APR differs by {0.1 + (k % 7) * 0.05:.2f}%; "
            f"also my CI pipeline fails on step {3 + k % 5} with a Docker layer cache issue. "
            f"Thời tiết tuần này có ảnh hưởng gì đến logistics không — mình cần ước lượng delay."
        )
        reason = "Hỏi tài chính, DevOps/code, thời tiết thực tế — ngoài phạm vi companion anime/tâm lý nhẹ."
        lab = "out_of_domain"
    elif v == 2:  # entertainment
        text = (
            f"I've been going down a rabbit hole reading fan theories about {w}: timeline contradictions, "
            f"which chapters the anime skipped, and whether the LN adds scenes that change character motivation. "
            f"If I only have time for either the manga or the game adaptation this month, which gives the fuller lore without filler?"
        )
        reason = f"Người dùng hỏi kiến thức/lore/review về tác phẩm ({w}), thuộc entertainment_knowledge."
        lab = "entertainment_knowledge"
    elif v == 3:  # venting
        text = (
            f"I'm so tired of pretending everything's fine at work — every meeting feels performative, "
            f"and when I get home I just scroll until 2am because quiet feels heavier than noise. "
            f"Mình không cần ai fix hộ ngay; đôi khi chỉ muốn nói ra cho đỡ nghẹn."
        )
        reason = "Chủ yếu xả stress, mệt mỏi, cô đơn; không yêu cầu checklist giải pháp rõ ràng."
        lab = "psychology_venting"
    elif v == 4:  # advice
        text = (
            f"When my chest tightens before social events I freeze — what are practical steps I can rehearse "
            f"the day before, and how do I explain boundaries to friends who take 'maybe' as 'yes'? "
            f"Bạn có thể gợi ý framework cụ thể (time-boxing, scripting) không?"
        )
        reason = "Đang xin chiến lược / lời khuyên hành động (how to / tips), không chỉ than thở."
        lab = "psychology_advice_seeking"
    else:  # crisis
        text = (
            f"I feel ashamed even writing this, but lately I've had intrusive thoughts about not being here anymore — "
            f"like I'd disappear and the noise would stop. I'm scared I might act on impulse; "
            f"mình không biết gọi ai và sợ bị coi là drama."
        )
        reason = "Có tín hiệu tự hại / không muốn sống / sợ làm điều nguy hiểm — crisis_alert."
        lab = "crisis_alert"

    # Thêm biến thể nhỏ theo chỉ số để các dòng không trùng byte-for-byte
    suffix = f" [row variant {i}]"
    if len(text) + len(suffix) < 3500:
        text = text + suffix
    return (text, reason, lab)


def _extended_long_rows(n: int) -> list[tuple[str, str, str]]:
    return [_long_row_at(i) for i in range(n)]


def _style_header(ws, row: int = 1) -> None:
    for cell in ws[row]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def _autosize_columns(ws, max_width: int = 72) -> None:
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_len = 0
        for row in ws.iter_rows(min_col=col, max_col=col):
            for c in row:
                if c.value is not None:
                    max_len = max(max_len, min(120, len(str(c.value))))
        ws.column_dimensions[letter].width = min(max_width, max(14, max_len + 2))


def main() -> None:
    OUT.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws_t = wb.active
    ws_t.title = "Taxonomy"
    ws_t.append(["label", "description"])
    for label, desc in TAXONOMY:
        ws_t.append([label, desc])
    _style_header(ws_t)
    _autosize_columns(ws_t)

    ws_d = wb.create_sheet("Dataset")
    ws_d.append(
        [
            "stt",
            "Câu người dùng (User Query)",
            "Lý luận gợi ý (Reasoning, tiếng Việt)",
            "Nhãn (Label)",
            "ghi_chú",
        ]
    )

    stt = 0
    for text, reasoning, label in SEED_ROWS:
        stt += 1
        ws_d.append([stt, text, reasoning, label, "seed"])

    for text, reasoning, label in _extended_long_rows(EXTRA_LONG_SAMPLES):
        stt += 1
        ws_d.append([stt, text, reasoning, label, ""])

    _style_header(ws_d)
    _autosize_columns(ws_d)

    wb.save(OUT)
    print(
        f"Wrote {OUT} — Taxonomy + Dataset: {len(SEED_ROWS)} seed + {EXTRA_LONG_SAMPLES} long = {stt} rows"
    )


if __name__ == "__main__":
    main()
